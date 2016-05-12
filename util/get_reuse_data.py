# coding=utf-8
from openpyxl import load_workbook
import argparse

#segement_xlsx Sheet3 必须是有无漏洞代码段和原函数表
#reuse_xlsx 第二列空出来
def get_reuse_data(segement_xlsx, reuse_xlsx):
	seg_ws = load_workbook(segement_xlsx)[u'Sheet3']
	segements = {}
	for row in seg_ws.rows:
		segements[row[0].value] = row[2].value
	
	reuse_wb = load_workbook(reuse_xlsx)
	reuse_ws = reuse_wb.active
	for row in reuse_ws.rows:
		try:
			row[1].value = segements[row[0].value]
		except Exception,e:
			print e
	
	reuse_wb.save(reuse_xlsx)

if __name__ == "__main__":
	parse = argparse.ArgumentParser()
	parse.add_argument("seg_xlsx", help="segement xlsx file path")
	parse.add_argument("reuse_xlsx", help="reuse xlsx file path")
	args = parse.parse_args()
	
	get_reuse_data(args.seg_xlsx, args.reuse_xlsx)
