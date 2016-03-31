#coding=utf-8
'''
Created on 2016年3月31日
@author: Bert
'''
import sys
sys.path.append("..")

from algorithm.ast import get_function_node
from openpyxl import load_workbook
from py2neo import Graph

if __name__ == "__main__":
    
    db = Graph("http://127.0.0.1/db/data/")
    #wireshark diff
    ws1 = load_workbook("Wireshark.xlsx")['Sheet3']
    for row in ws1.rows:
        try:
            n = get_function_node(db, row[0].value)
        except Exception as e:
            try:
                n = get_function_node(db, row[2].value)
            except:
                print "%s and %s not found" % (row[0].value, row[2].value)
    
    #wireshark patch Function
    for row in ws1.rows:
        try:
            n = get_function_node(db, row[1].value)
        except Exception as e:
            print "%s not found" % row[1].value
    
    #ffmpeg diff
    ws2= load_workbook("ffmpeg.xlsx")['Sheet3']
    for row in ws2.rows:
        try:
            n = get_function_node(db, row[0].value)
        except Exception as e:
            try:
                n = get_function_node(db, row[2].value)
            except:
                print "%s and %s not found" % (row[0].value, row[2].value)
    
    #ffmpeg patch Function
    for row in ws2.rows:
        try:
            n = get_function_node(db, row[1].value)
        except Exception as e:
            print "%s not found" % row[1].value
    
    # code reuse
    wb = load_workbook("data.xlsx")
    ws3 = wb['Sheet1']
    for row in ws3.rows:
        try:
            n = get_function_node(db, row[0].value)
        except Exception as e:
            print "%s not found" % row[0].value
    
    ws4 = wb['Sheet2']
    for row in ws4.rows:
        try:
            n = get_function_node(db, row[0].value)
        except Exception as e:
            print "%s not found" % row[0].value