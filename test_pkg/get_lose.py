#!/usr/bin/python
# -*- coding:utf-8 -*-

from openpyxl import load_workbook, Workbook

def get_lose(source_xlsx, result_xlsx):
	ret = load_workbook(result_xlsx).active
	result_list = []
	for row in ret.rows:
		result_list.append(row[0].value)

	lose = []
	src = load_workbook(source_xlsx)['Sheet3']
	for row in src.rows:
		if row[0].value in result_list:
			continue
		else:
			lose.append((row[0].value,row[1].value,row[2].value))
	return lose

if __name__ == "__main__":
	wb = Workbook()
	
	ffmpeg = wb.create_sheet("ffmpeg", 0)
	#get ffmpeg lose
	ret =  get_lose("ffmpeg.xlsx", "ffmpeg_diff.xlsx")
	for r in ret:
		ffmpeg.append(r)

	wireshark = wb.create_sheet("wireshark", 1)
	#get wiresgark lose
	ret = get_lose("wireshark.xlsx", "wireshark_diff.xlsx")
	for r in ret:
		wireshark.append(r)

	linux = wb.create_sheet("linux", 2)	
	#get linux lose
	ret = get_lose("linux.xlsx", "linux_diff.xlsx")
	for r in ret:
		linux.append(r)

	wb.save("lose.xlsx")