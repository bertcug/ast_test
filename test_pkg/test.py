# coding=utf-8
'''
Created on 2016年3月28日

@author: Bert
'''
import sys
sys.path.append("..")
import traceback
import datetime
from algorithm.ast import serializedAST, get_function_ast_root,get_function_node_by_ast_root
from openpyxl import load_workbook, Workbook
from algorithm.suffixtree import suffixtree
from py2neo import Graph
from segement_comp import get_type_mapping_table
import sqlite3
import argparse

def search_vuln_seg_in_func(db1, vuln_seg, vuln_func, var_map, db2, func_name, suffix_obj):
    
    print "[%s] processing %s VS %s" % (
                                   datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S"),
                                   vuln_seg, func_name)
    
    
    vuln_seg_func = get_function_ast_root(db1, vuln_seg)
    if vuln_seg_func is None:
        vuln_seg_func = get_function_ast_root(db1, vuln_func)
        
    if vuln_seg_func is None:
        print "%s  %s not found" % (vuln_seg, vuln_func)
        return (vuln_seg+"-"+vuln_func, func_name, "vuln_not_found")
    
    patched_func = get_function_ast_root(db2, func_name)
    if patched_func is None:
        print "%s is not found" % func_name
        return (vuln_seg, func_name, "patch_not_found")
    
    o1 = serializedAST(db1)
    o1.variable_maps = var_map
    ret = o1.genSerilizedAST(vuln_seg_func)
    
    #delete FunctionDef and CompoundStatement node
    pattern1 = ";".join(ret[0][2:])
    pattern2 = ";".join(ret[1][2:]) 
    pattern3 = ";".join(ret[2][2:])
    pattern4 = ";".join(ret[3][2:])
    pattern5 = ";".join(ret[4][2:])
    
    tmp = serializedAST(db2).genSerilizedAST(patched_func)
    s1 = ";".join(tmp[0][2:])
    s2 = ";".join(tmp[1][2:])
    s3 = ";".join(tmp[2][2:])
    s4 = ";".join(tmp[3][2:])
    s5 = ";".join(tmp[4][2:])
      
    report = {}
    if suffix_obj.search(s1, pattern1):
        report['distinct_type_and_const'] = True
    else:
        report['distinct_type_and_const'] = False
        
    if suffix_obj.search(s2, pattern2):
        report['distinct_const_no_type'] = True
    else:
        report['distinct_const_no_type'] = False
        
    if suffix_obj.search(s3, pattern3):
        report['distinct_type_no_const'] = True
    else:
        report['distinct_type_no_const'] = False
        
    if suffix_obj.search(s4, pattern4):
        report['no_type_no_const'] = True
    else:
        report['no_type_no_const'] = False

    if suffix_obj.search(s5, pattern5):
        report['no_mapping'] = True
    else:
        report['no_mapping'] = False
    
    #begin cfg
#     patch_root = get_function_node_by_ast_root(db2, patched_func)
#     vuln_seg_root = get_function_node_by_ast_root(db1, vuln_seg_func)
#     match, simi = func_cfg_similarity(patch_root, db2, vuln_seg_root, db1)
    
    return (vuln_seg, func_name, "success", report["distinct_type_and_const"],
                       report["distinct_const_no_type"], report["distinct_type_no_const"],
                       report["no_type_no_const"], report['no_mapping'])
    
def wireshark_diff():
    data = load_workbook("/home/bert/Documents/data/wireshark.xlsx", read_only=True)[u'Sheet3']
    suffix_obj = suffixtree()
    
    wb = Workbook()
    ws = wb.active
    
    db1 = Graph("http://127.0.0.1:7473/db/data/")
    db2 = Graph()
    
    for row in data.rows:
        vuln_seg = row[0].value
        patched_name = vuln_seg[:14] + "PATCHED_" + row[2].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[2].value
        var_map = get_type_mapping_table(db2, vuln_name)
        try:
            
            ret = search_vuln_seg_in_func(db1, vuln_seg, row[2].value,var_map, db2, patched_name, suffix_obj)
            ws.append(ret)
        except Exception as e:
            print e
            ws.append((vuln_seg, patched_name, "failed"))
        
        wb.save("/home/bert/Documents/data/wireshark_diff.xlsx")
        
    suffix_obj.close()
    print "wireshark all works done"
    
def ffmpeg_diff():
    data = load_workbook("/home/bert/Documents/data/ffmpeg.xlsx", read_only=True)[u'Sheet3']
    suffix_obj = suffixtree()
    
    wb = Workbook()
    ws = wb.active
    
    db1 = Graph("http://127.0.0.1:7473/db/data/")
    db2 = Graph()
    
    for row in data.rows:
        vuln_seg = row[0].value
        patched_name = vuln_seg[:14] + "PATCHED_" + row[2].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[2].value
        var_map = get_type_mapping_table(db2, vuln_name)
        try:
            ret = search_vuln_seg_in_func(db1, vuln_seg, row[2].value, var_map, db2, patched_name, suffix_obj)
            ws.append(ret)
            
        except Exception as e:
            print e
            ws.append((vuln_seg, patched_name, "failed"))
        wb.save("/home/bert/Documents/data/ffmpeg_diff.xlsx")
    
    suffix_obj.close()
    print "ffmpeg all works done"
    
def linux_diff():
    data = load_workbook("/home/bert/Documents/data/linux.xlsx", read_only=True)[u'Sheet3']
    suffix_obj = suffixtree()
    
    wb = Workbook()
    ws = wb.active
    
    db1 = Graph("http://127.0.0.1:7473/db/data/")
    db2 = Graph()
    
    for row in data.rows:
        vuln_seg = row[0].value
        patched_name = vuln_seg[:14] + "PATCHED_" + row[2].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[2].value
        var_map = get_type_mapping_table(db2, vuln_name)
        try:
            ret = search_vuln_seg_in_func(db1, vuln_seg, row[2].value, var_map, db2, patched_name, suffix_obj)
            ws.append(ret)
        except Exception as e:
            print e
            ws.append((vuln_seg, patched_name, "failed"))
        wb.save("/home/bert/Documents/data/linux_diff.xlsx")
    
    suffix_obj.close()
    print "linux all works done"

def firefox_diff():
    data = load_workbook("/home/bert/Documents/data/firefox.xlsx", read_only=True)[u'Sheet3']
    suffix_obj = suffixtree()
    
    wb = Workbook()
    ws = wb.active
    
    db = Graph("http://127.0.0.1:7474/db/data/")
    
    for row in data.rows:
        vuln_seg = row[0].value
        patched_name = vuln_seg[:14] + "PATCHED_" + row[2].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[2].value
        var_map = get_type_mapping_table(db, vuln_name)
        try:
            ret = search_vuln_seg_in_func(db, vuln_seg, row[2].value, var_map, db, patched_name, suffix_obj)
            ws.append(ret)
        except Exception as e:
            print e
            ws.append((vuln_seg, patched_name, "failed"))
        wb.save("/home/bert/Documents/data/firefox_diff.xlsx")
    
    suffix_obj.close()
    print "firefox all works done"

def get_segements(segements, patch_func_name):
    cveid = patch_func_name[0:13]
    ret = []
    for segement in segements:
        if segement[0].startswith(cveid):
            ret.append(segement)
    return ret    

def code_reuse(table_name, worksheet):
    result_db = sqlite3.connect("/home/bert/Documents/data/code_reuse.db")
    result_db.execute('''create table if not exists %s(
        vuln_segement CHAR(50) NOT NULL,
        reuse_func CHAR(50) NOT NULL,
        status CHAR(10) NOT NULL,
        distinct_type_and_const BOOLEAN,
        distinct_const_no_type BOOLEAN,
        distinct_type_no_const BOOLEAN,
        no_type_no_const BOOLEAN)
    ''' % table_name)
       
    db1 = Graph("http://127.0.0.1:7473/db/data/")
    db2 = Graph()
    suffix_obj = suffixtree()
     
    for row in worksheet.rows:
        #check
        ret = result_db.execute("select * from %s where vuln_segement=? and reuse_func=?" % table_name, (row[0].value, row[2].value))
        if ret.fetchone():
            continue
        
        vuln_seg = row[0].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[1].value
        
        try:
            var_map = get_type_mapping_table(db2, vuln_name)
            ret = search_vuln_seg_in_func(db1, row[0].value, row[1].value, var_map, db1, row[2].value, suffix_obj)
            
            if ret[2] == "success":
                result_db.execute("insert into %s values(?,?,?,?,?,?,?)" % table_name, ret)
            else:
                result_db.execute("insert into %s(vuln_segement, reuse_func, status) values(?,?,?)" % table_name, ret)
            result_db.commit()
        except Exception as e:
            result_db.execute("insert into %s(vuln_segement, reuse_func, status) values(?,?,?)" % table_name,
                               (row[0].value, row[2].value, "failed") ) 
            print e

def firefox_code_reuse(table_name):
    result_db = sqlite3.connect("/home/bert/Documents/data/firefox_code_reuse.db")
    result_db.execute('''create table if not exists %s(
        vuln_segement CHAR(50) NOT NULL,
        reuse_func CHAR(50) NOT NULL,
        status CHAR(10) NOT NULL,
        distinct_type_and_const BOOLEAN,
        distinct_const_no_type BOOLEAN,
        distinct_type_no_const BOOLEAN,
        no_type_no_const BOOLEAN,
        no_mapping BOOLEAN)
    ''' % table_name)
       
    db = Graph("http://127.0.0.1:7474/db/data/")
    suffix_obj = suffixtree()
    
    worksheet = load_workbook("/home/bert/Documents/data/firefox_reuse.xlsx").active
    for row in worksheet.rows:
        #check
        ret = result_db.execute("select * from %s where vuln_segement=? and reuse_func=?" % table_name, (row[0].value, row[2].value))
        if ret.fetchone():
            continue
        
        vuln_seg = row[0].value
        vuln_name = vuln_seg[:14] + "VULN_" + row[1].value
        
        try:
            var_map = get_type_mapping_table(db, vuln_name)
            ret = search_vuln_seg_in_func(db, row[0].value, row[1].value, var_map, db, row[2].value, suffix_obj)
            
            if ret[2] == "success":
                result_db.execute("insert into %s values(?,?,?,?,?,?,?,?)" % table_name, ret)
            else:
                result_db.execute("insert into %s(vuln_segement, reuse_func, status) values(?,?,?)" % table_name, ret)
            result_db.commit()
        except Exception as e:
            result_db.execute("insert into %s(vuln_segement, reuse_func, status) values(?,?,?)" % table_name,
                               (row[0].value, row[2].value, "failed") ) 
            print e

    print "firefox reuse works done"

if __name__ == "__main__":
    # parse = argparse.ArgumentParser()
    # group = parse.add_mutually_exclusive_group()
    # group.add_argument("-diff", action="store_true", help="test diff")
    # group.add_argument("-reuse", action="store_true", help="test reuse")

    # parse.add_argument("soft", help="software name, used as table_name or excel")

    # parse.add_argument("-seg_xlsx",help="segement xlsx path, use Sheet3")
    # parse.add_argument("-seg_port", type=int, help="segement neo4j db port")
    # pase.add_argument("-vp_port", type=int, help="vuln patched func's neo4j port")

    arg = sys.argv[1]
    if arg == "wireshark":
        wireshark_diff()
    elif arg == "ffmpeg":
        ffmpeg_diff()
    elif arg == "reuse":
        
        print "wireshark code reuse"
        ws = load_workbook("/home/bert/Documents/data/wireshark_reuse.xlsx").active
        code_reuse("wireshark", ws)
        
        print "ffmpeg code reuse"
        ws = load_workbook("/home/bert/Documents/data/ffmpeg_reuse.xlsx").active
        code_reuse("ffmpeg", ws)
        
        print "linux code reuse"
        ws = load_workbook("/home/bert/Documents/data/linux_reuse.xlsx").active
        code_reuse("linux", ws)
        
        print "all works done!"
    elif arg == "linux":
        linux_diff()
    elif arg == "firefox_diff":
        firefox_diff()
    elif arg == "firefox_reuse":
        firefox_code_reuse("firefox")
    else:
        print "argument error"
    
    