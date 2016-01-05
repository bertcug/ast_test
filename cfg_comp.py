#coding=utf-8
'''
Created on Jan 5, 2016
@author: bert
'''

from openpyxl import load_workbook, Workbook
from algorithm.ast import get_function_node
from algorithm.graph import func_cfg_similarity
import time
import py2neo
import multiprocessing
from multiprocessing import Pool
from db.models import vulnerability_info, cve_infos, get_connection
from algorithm.ast import get_function_node
import datetime

def func_cfg_similarity_process(vuln_info, lock):
    conn = get_connection()
    neo4jdb = py2neo.Graph()
     
    start_time = time.time()
    cve_info = vuln_info.get_cve_info(conn)
    soft = cve_info.get_soft(conn)
    conn.close()
    
    print "[%s] processing %s" % (datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S"),
                                   cve_info.cveid)
    
    vuln_name = cve_info.cveid.replace(u"-", u"_").upper() + u"_VULN_" + vuln_info.vuln_func
    patch_name = cve_info.cveid.replace(u"-", u"_").upper() + u"_PATCHED_" + vuln_info.vuln_func
    
    vuln_func = get_function_node(neo4jdb, vuln_name)
    if vuln_func is None:
        lock.acquire()
        
        wb = load_workbook("cfg_result.xlsx")
        ws = wb.active
        line = (cve_info.cveid, soft.software_name + "-" + soft.software_version,
                 vuln_info.vuln_func, vuln_info.vuln_file[41:], 
                "vuln_func_not_found", 0.00, 0)
        ws.append(line)
        wb.save("cfg_result.xlsx")
        
        lock.release()
        return
         
    patch_func = get_function_node(neo4jdb, patch_name)
    if patch_name is None:
        lock.acquire()
        
        wb = load_workbook("cfg_result.xlsx")
        ws = wb.active
        line = (cve_info.cveid, soft.software_name + "-" + soft.software_version,
                vuln_info.vuln_func, vuln_info.vuln_file[41:], 
                "patch_func_not_found", 0.00, 0)
        ws.append(line)
        wb.save("cfg_result.xlsx")
        
        lock.release()
        return
    
    match, simi = func_cfg_similarity(vuln_func, neo4jdb, patch_func, neo4jdb)
   
    #u"success"
    end_time = time.time()
    cost = round(end_time - start_time, 2)
    
    lock.acquire()
        
    wb = load_workbook("cfg_result.xlsx")
    ws = wb.active
    line = (cve_info.cveid, soft.software_name + "-" + soft.software_version,
            vuln_info.vuln_func, vuln_info.vuln_file[41:], 
                match, simi, cost)
    ws.append(line)
    wb.save("cfg_result.xlsx")
       
    lock.release()
    
if __name__ == "__main__":
    db_conn = get_connection()
    if db_conn is None:
        print u"数据库连接失败"
        exit(0)
    
    cur = db_conn.cursor()
    cur.execute("select * from vulnerability_info")
    rets = cur.fetchall()
    cur.close()
    infos = []
    for ret in rets:
        soft = vulnerability_info(ret).get_cve_info(db_conn).get_soft(db_conn)
        if soft.software_name == "ffmpeg":
            infos.append(ret)
 
    wb = Workbook()
    ws = wb.active
    ws.title = u"CFG测试结果"
    header = [u'CVE编号', u"软件版本", u"漏洞函数", u"漏洞文件","Match","Similarity", "cost"]
    ws.append(header)
    wb.save("cfg_result.xlsx")
    
    pool = Pool(processes = 10)
    lock = multiprocessing.Manager().Lock()
    for info in infos:
        pool.apply(func_cfg_similarity_process, (vulnerability_info(info),lock))
    
    pool.close()
    pool.join()
    
    print "all works done!"