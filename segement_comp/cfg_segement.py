#coding=utf-8
'''
Created on 2016年1月29日
@author: bert
'''
import sys
sys.path.append("..")

from openpyxl import load_workbook, Workbook
from algorithm.ast import get_function_node
from algorithm.graph import func_cfg_similarity
import time
import py2neo
import datetime

def segement_cfg_similarity_process(vuln_name, patch_name, neo4jdb, worksheet):
    start_time = time.time()
    print "[%s] processing %s" % (datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S"),
                                   vuln_name + " vs " + patch_name)
    
    
    #检查数据库里面是否可以找到该函数，找不到相似度为0
    vuln_func = get_function_node(neo4jdb, vuln_name)
    if vuln_func is None:
        line = (vuln_name, patch_name, "vuln_func_not_found", 0.00, 0)
        worksheet.append(line)
        return
    
    #检查数据库里面是否可以找到该函数，找不到相似度为0     
    patch_func = get_function_node(neo4jdb, patch_name)
    if patch_func is None:
        line = (vuln_name, patch_name, "patch_func_not_found", 0.00, 0)
        worksheet.append(line)
        return
    
    #计算是否匹配和相似度
    match, simi = func_cfg_similarity(vuln_func, neo4jdb, patch_func, neo4jdb)
   
    #u"success"
    end_time = time.time()
    cost = round(end_time - start_time, 2)
    
    line = (vuln_name, patch_name, match, simi, cost)
    worksheet.append(line)

if __name__ == "__main__":
    try:
        neo4jdb = py2neo.Graph("http://127.0.0.1:7475/db/data/")
    except Exception:
        print "数据库连接失败"
        
    wb = load_workbook("test3.xlsx", read_only=True)
    ws = wb[u'Sheet3']
    
    workbook = Workbook()
    worksheet = wb.active
    worksheet.title = u"CFG代码段测试结果"
    header = [u'漏洞段', u"无漏洞段", u"是否匹配", u"相似度", u"耗时"]
    worksheet.append(header)
    
    '''
    for row in ws.rows:
        if row[0].value == u"漏洞函数名":
            continue
        
        try:
            segement_cfg_similarity_process(row[0].value, row[1].value, neo4jdb, worksheet)
        except Exception:
            print "process " + row[0].value + "error"
    '''
    segement_cfg_similarity_process("CVE_2014_8544_VULN_COMPLETE_0", "CVE_2014_8544_PATCHED_COMPLETE_0", neo4jdb, worksheet)
    workbook.save("cfg_segement_result.xlsx")