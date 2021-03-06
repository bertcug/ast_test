#coding=utf-8
import sys
sys.path.append("..")

from openpyxl import load_workbook, Workbook
from algorithm.ast import get_function_node
from algorithm.ast import serializedAST, get_function_ast_root
import time
import py2neo
import datetime
import re
from algorithm.suffixtree import suffixtree
from segement_comp import get_type_mapping_table

def segement_ast_similarity_process(vuln_name, patch_name, neo4jdb, org_func_name, type_mapping,
                                     worksheet, suffix_tree_obj):
    start_time = time.time()
    print "[%s] processing %s" % (datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S"),
                                   vuln_name + " vs " + patch_name)
    
    #检查数据库里面是否可以找到该函数
    vuln_func = get_function_ast_root(neo4jdb, vuln_name)
    if vuln_func is None:
        line = (vuln_name, patch_name, "vuln_func_not_found", "-", "-", "-", "-", 0, 
                org_func_name, type_mapping.__str__())
        worksheet.append(line)
        return
     
    #检查数据库里面是否可以找到该函数    
    patched_func = get_function_ast_root(neo4jdb, patch_name)
    if patched_func is None:
        line = (vuln_name, patch_name, "patch_func_not_found", "-", "-", "-", "-", 0, 
                org_func_name, type_mapping.__str__())
        worksheet.append(line)
        return
    
    
    
    #序列化AST返回值是一个数组，0元素是序列化的AST字符串，1元素是节点个数，AST字符串以;结尾，需要去掉结尾的;
    o1 = serializedAST(neo4jdb)
    o1.data_type_mapping = type_mapping
    
    ret = o1.genSerilizedAST(vuln_func)
    #delete FunctionDef and CompoundStatement node [2:]
    pattern1 = ";".join(ret[0][2:])
    pattern2 = ";".join(ret[1][2:])
    pattern3 = ";".join(ret[2][2:])
    pattern4 = ";".join(ret[3][2:])
    
    tmp = o1.genSerilizedAST(patched_func)
    s1 = ";".join(tmp[0])
    s2 = ";".join(tmp[0])
    s3 = ";".join(tmp[0])
    s4 = ";".join(tmp[0])
    
    report = {}
    if suffix_tree_obj.search(s1, pattern1):
        report['distinct_type_and_const'] = True
    else:
        report['distinct_type_and_const'] = False
        
    if suffix_tree_obj.search(s2, pattern2):
        report['distinct_const_no_type'] = True
    else:
        report['distinct_const_no_type'] = False
        
    if suffix_tree_obj.search(s3, pattern3):
        report['distinct_type_no_const'] = True
    else:
        report['distinct_type_no_const'] = False
        
    if suffix_tree_obj.search(s4, pattern4):
        report['distinct_type_no_const'] = True
    else:
        report['no_type_no_const'] = False
    
    end_time = time.time()
    cost = round(end_time - start_time, 2)
    line = (vuln_name, patch_name,"success", report['distinct_type_and_const'],
            report['distinct_const_no_type'], report['distinct_type_no_const'],
            report['distinct_type_no_const'], cost, org_func_name, type_mapping.__str__())
    
    worksheet.append(line)

if __name__ == "__main__":
    try:
        neo4jdb = py2neo.Graph("http://127.0.0.1:7499/db/data/")
    except Exception:
        print u"数据库连接失败:7499"
    
    try:
        org_db = py2neo.Graph()
    except Exception:
        print u"数据库连接失败:7474"
        
    wb = load_workbook("test3.xlsx", read_only=True)
    ws = wb[u'Sheet3']
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = u"AST代码段测试结果"
    header = [u'漏洞段', u"无漏洞段", u"计算状态", u"区分类型和常量",u"区分常量不区分类型",u"区分类型不区分常量",u"不区分常量和类型", 
              u"耗时", u"原漏洞函数", u"类型映射"]
    worksheet.append(header)
    
    suffix_tree_obj = suffixtree()
    for row in ws.rows:
        type_mapping = {'other':'v'}
        if row[2].value != 0:
            func_name = row[0].value[:19] + row[2].value
            type_mapping = get_type_mapping_table(org_db, func_name)    
        try:
            segement_ast_similarity_process(row[0].value, row[1].value, neo4jdb, row[2].value,
                                            type_mapping, worksheet, suffix_tree_obj)
            workbook.save("ast_segement_result.xlsx")
        except Exception as e:
            print "process " + row[0].value + "error"
            print e
    
    
    suffix_tree_obj.close()
    print "all works done!"
    
    