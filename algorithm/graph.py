# coding=utf-8
'''
Created on 2015年11月9日

@author: Bert
'''
from igraph import Graph
import time
import py2neo
from db.models import vulnerability_info, cve_infos, get_connection
from algorithm.ast import get_function_node
from openpyxl import load_workbook

def get_cfg_nodes(neo4j_db, function_node):
    query = "match (n {functionId:%d, isCFGNode:'True'} return n"
    records = neo4j_db.cypher.execute(query)
   
    nodes = []
    for record in records:
        nodes.append(record[0])
    
    return nodes

def get_cfg_edges(neo4j_db, function_node):
    query = "match (n {functionId:%d, isCFGNode:'True'})-[e:`FLOWS_TO`]->(m) return e"\
             % function_node.properties['functionId']
    records = neo4j_db.cypher.execute(query)
    
    edges = []
    for record in records:
        edges.append(record[0])
    
    return edges


def get_ddg_edges(neo4j_db, function_node):
    query = "match(n {functionId:%d, isCFGNode:'True'})-[e:`REACHES`]->(m) return e"\
             % function_node.properties['functionId']
    records = neo4j_db.cypher.execute(query)
    
    edges = []
    for record in records:
        edges.append(record[0])
    
    return edges

def get_cdg_edges(neo4j_db, function_node):
    query = "match(n {functionId:%d, isCFGNode:'True'})-[e:`CONTROLS`]->(m) return e"\
             % function_node.properties['functionId']
    records = neo4j_db.cypher.execute(query)
    
    edges = []
    for record in records:
        edges.append(record[0])
    
    return edges

def get_func_file(neo4j_db, function_node):
    query = "start n=node(%d) match (m {type:'File'})-[:`IS_FILE_OF`]->(n) return m.filepath" % function_node._id
    records = neo4j_db.cypher.execute(query)
    return records.one

def translate_cfg(neo4j_db, function_node):
    cfg_nodes = get_cfg_nodes(neo4j_db, function_node)
    cfg_edges = get_cfg_edges(neo4j_db, function_node)
    
    #create igraph cfg
    g = Graph(directed = True)
    #add node and node properties
    for cfgNode in cfg_nodes :
        node_prop = {'code':cfgNode.properties['code'],'type':cfgNode.properties['type']}
        g.add_vertex(str(cfgNode._id),**node_prop)
    #add edge and edge properties
    for cfgEdge in cfg_edges :
        startNode = str(cfgEdge.start_node._id)
        endNode = str(cfgEdge.end_node._id)
        edge_prop = {'flowLabel':cfgEdge.properties['flowLabel']}
        g.add_edge(startNode,endNode,**edge_prop)
    return g

def node_compat_fn(g1,g2,n1,n2):
    if g1.vs[n1]['type']==g2.vs[n2]['type'] :
        return True
    else:
        return False

def edge_compat_fn(g1,g2,e1,e2):
    if g1.es[e1]['flowLabel'] == g2.es[e2]['flowLabel'] :
        return True
    else:
        return False
    
def cal_similarity(srcCFG,tarCFG,vertexMap):
    count = 0
    sum = 0
    if vertexMap:
        sum = len(vertexMap)
    for i in range(sum) :
        if srcCFG.vs[vertexMap[i]]['code'] == tarCFG.vs[i]['code'] :
            count +=1
    return round((float(count)/float(sum)), 2)

def func_cfg_similarity(func1, db1, func2, db2):
    srcCFG = translate_cfg(db1, func1)
    targetCFG = translate_cfg(db2, func2)
    ret = srcCFG.get_subisomorphisms_vf2(other = targetCFG,node_compat_fn = node_compat_fn,
                                         edge_compat_fn = edge_compat_fn)
    if len(ret) == 0:
        return False, 0
    else:
        rs = []
        for r in ret:
            rs.append(cal_similarity(srcCFG, targetCFG, r))
        
        return True, round(max(rs), 4)

def func_cfg_similarity_process(vuln_info, lock):
    conn = get_connection()
    neo4jdb = py2neo.Graph()
     
    start_time = time.time()
    cve_info = vuln_info.get_cve_info()
    conn.close()
    
    vuln_name = cve_info.cveid.replace(u"-", u"_").upper() + u"_VULN_" + vuln_info.vuln_func
    patch_name = cve_info.cveid.replace(u"-", u"_").upper() + u"_PATCHED_" + vuln_info.vuln_func
    
    vuln_func = get_function_node(vuln_name, neo4jdb)
    if vuln_func is None:
        lock.acquire()
        
        wb = load_workbook("cfg_result.xlsx")
        ws = wb.active
        line = (cve_info.cveid, vuln_info.vuln_func, vuln_info.vuln_file[41:], 
                "vuln_func_not_found", 0.00, 0)
        ws.append(line)
        wb.save("cfg_result.xlsx")
        
        lock.release()
        return
         
    patch_func = get_function_node(patch_name, neo4jdb)
    if patch_name is None:
        lock.acquire()
        
        wb = load_workbook("cfg_result.xlsx")
        ws = wb.active
        line = (cve_info.cveid, vuln_info.vuln_func, vuln_info.vuln_file[41:], 
                "patch_func_not_found", 0.00, 0)
        ws.append(line)
        wb.save("cfg_result.xlsx")
        
        lock.release()
        return
    
    match, simi = func_cfg_similarity(vuln_func, neo4jdb, patch_func, neo4jdb)
   
    u"success"
    end_time = time.time()
    cost = round(end_time - start_time, 2)  
    
